import openpyxl as xl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


class ExcelReadWriter(object):
    def __init__(self):
        self.workbook = None

    def load(self, path, data_only=False, read_only=True):
        self.workbook = xl.load_workbook(
            path, data_only=data_only, read_only=read_only)

    def write_to_new_file(self, path):
        write_work_book = xl.Workbook(write_only=True)
        ws = write_work_book.create_sheet('Sample list')
        read_work_sheet = self.workbook['Sample list']
        normal_style = self.workbook._named_styles['Normal']
        for row in read_work_sheet.iter_rows():
            write_row = list()
            for cell in row:
                write_cell = WriteOnlyCell(ws, value=cell.value)
                if cell.font != normal_style.font:
                    write_cell.font = cell.font
                if cell.fill is not None:
                    write_cell.fill = cell.fill
                write_row.append(write_cell)
            ws.append(write_row)

        write_work_book.save(path)

    def write_with_dual_mode(self, path):
        self.workbook.save(path)
