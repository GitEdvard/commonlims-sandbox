from __future__ import absolute_import
import openpyxl as xl
import os
from six import BytesIO
from pprint import pprint
from sentry.testutils import TestCase
from sentry.models.file import File
from clims.models.file import OrganizationFile
from clims.models.substance import Substance
from clims.handlers import HandlerContext
from commonlims.utility.gemstone_sample import GemstoneSample
from commonlims.utility.test_utils import create_plugin
from commonlims.test.resources.resource_bag import gemstone_xlsx_path
from commonlims.test.resources.resource_bag import read_gemstone_xlsx
from commonlims.utility.submission_handler import GemstoneSubmissionHandler


class TestSampleSubmissionXlsx(TestCase):
    def setUp(self):
        plugin = create_plugin(self.organization)
        self.register_extensible(GemstoneSample, plugin)
        self.handler_context = HandlerContext(self.organization)

    def _create_xlsx_organization_file(self):
        name = os.path.basename(gemstone_xlsx_path())
        file_model = File.objects.create(
            name=name,
            type='substance-batch-file',
            headers=list(),
        )
        contents = read_gemstone_xlsx()
        file_like_obj = BytesIO(contents)
        file_model.putfile(file_like_obj)
        return OrganizationFile(name=name, organization=self.organization, file=file_model)

    def _as_rows(self, orgfile):
        """
        Returns the file as an excel file
        """
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile(suffix='.xlsx') as tmp:
            with open(tmp.name, 'wb') as f:
                for _, chunk in enumerate(orgfile.getfile()):
                    f.write(chunk)

            workbook = xl.load_workbook(tmp.name)
        sample_list_sheet = workbook['Samples']
        name_index = 1
        preciousness_index = 4
        color_index = 5
        rows = []
        for row in sample_list_sheet.iter_rows(min_row=4):
            sample_name = str(row[name_index].value)
            preciousness = str(row[preciousness_index].value)
            color = str(row[color_index].value)
            line = ','.join([sample_name, preciousness, color])
            rows.append(line)
        return rows

    def test_import_xlsx_here(self):
        workbook = xl.load_workbook(gemstone_xlsx_path())
        sample_list_sheet = workbook['Samples']
        name_index = 1
        preciousness_index = 4
        color_index = 5
        rows = []
        for row in sample_list_sheet.iter_rows(min_row=3):
            sample_name = str(row[name_index].value)
            preciousness = str(row[preciousness_index].value)
            color = str(row[color_index].value)
            line = ','.join([sample_name, preciousness, color])
            rows.append(line)

        pprint(rows)
        assert 1 == 1

    def test_import_xlsx_here2(self):
        workbook = xl.load_workbook(gemstone_xlsx_path(), data_only=True)
        sample_list_sheet = workbook['Samples']
        rows = []
        for row in sample_list_sheet.iter_rows(min_row=3):
            line_contents = [str(cell.value) for cell in row]
            line = ','.join(line_contents)
            rows.append(line)
        contents = '\n'.join(rows)
        print(contents)
        assert 1 == 1

    def test_extract_excel_file_from_organization_file(self):
        orgfile = self._create_xlsx_organization_file()
        rows = self._as_rows(orgfile.file)
        pprint(rows)
        assert 1 == 1

    def test_gemstone_submission_handler__with_xlsx__3_samples_found(self):
        file = self._create_xlsx_organization_file()
        handler = GemstoneSubmissionHandler(context=self.handler_context, app=self.app)
        handler.handle(file)
        all_samples = Substance.objects.all()
        all_sample_names = [sample.name for sample in all_samples]
        expected_sample_names = [
            'Sample-Frink1',
            'Sample-Frink2',
            'Sample-Frink3',
        ]
        assert set(expected_sample_names).issubset(set(all_sample_names))
